#!/usr/bin/env python3
"""
停车场系统助手 - 整合API操作与本地Excel同步

功能：
1. 修改车牌 → API修改 + 同步更新本地表格
2. 新增人员+车牌 → API创建用户+添加凭证+开通月租 + 同步本地表格
3. 黑白名单操作 → 仅API操作，不同步本地表格

表格路径: F:\ShareCache\智能化系统\门禁系统\2024政府年门禁、车牌审核汇总表.xlsx

使用方法：
  python parking_helper.py modify <姓名> <新车牌>
  python parking_helper.py add <单位> <姓名> <车牌> <电话> <套餐> <有效期> [备注]

示例：
  python parking_helper.py modify 张一 粤S22222
  python parking_helper.py add 卫生健康局 陈座娣 粤BF78K9 13480980227 地面月卡 2030-12-30 借调
"""

import hashlib
import requests
import urllib3
import json
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
import sys

urllib3.disable_warnings()

# ==================== 配置 ====================
BASE_URL = "https://10.0.12.1:9091"
ACCOUNT = "9999"
PASSWORD = "88888888"

# 本地表格路径 (Windows路径 -> WSL路径)
XLSX_PATH = "/mnt/f/ShareCache/智能化系统/门禁系统/2024政府年门禁、车牌审核汇总表.xlsx"

# Excel列索引 (1-based)
COL_UNIT = 1       # 单位名称
COL_TYPE = 2       # 类型
COL_WORK_ID = 3    # 工作证编号
COL_NAME = 4       # 姓名
COL_POSITION = 5   # 职务
COL_PLATE = 6      # 车牌号码
COL_PHONE = 7      # 联系电话
COL_REMARK = 8     # 备注

# 套餐ID映射
SEAL_MAP = {
    "地面月卡": "p220447822150",
    "地面临时车": "p220447822154",
    "地库临时车": "p220447822155",
    "地库月卡": "p220447822159",
}

# 车牌颜色判断
def get_plate_color(plate: str) -> int:
    """根据车牌号判断颜色"""
    if not plate or len(plate) < 2:
        return 3  # 默认蓝牌
    # 绿牌：粤B/D/F开头的新能源车
    if plate.startswith("粤B") or plate.startswith("粤D") or plate.startswith("粤F"):
        if len(plate) == 8:  # 新能源车牌8位
            return 5  # 绿牌
    return 3  # 蓝牌


# ==================== 停车场API ====================

def login():
    """登录获取Token"""
    pw_md5 = hashlib.md5(PASSWORD.encode()).hexdigest()
    r = requests.post(
        f"{BASE_URL}/api/systemcenter/auth/login",
        json={"account": ACCOUNT, "password": pw_md5},
        verify=False, timeout=10
    )
    data = r.json()
    if data.get('code') == 200:
        return data['data']['token']
    raise Exception(f"登录失败: {data}")

def get_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

def get_groups(token):
    """获取组织列表"""
    r = requests.get(
        f"{BASE_URL}/api/systemcenter/group",
        headers=get_headers(token),
        params={"pageIndex": 0, "pageSize": 100},
        verify=False, timeout=10
    )
    return r.json().get('data', {}).get('rows', [])

def find_group_id(token, group_name_hint: str):
    """根据名称提示找组织ID"""
    groups = get_groups(token)
    for g in groups:
        if group_name_hint in g.get('name', ''):
            return g['id'], g['name']
    return None, None

def get_new_person_no(token):
    """获取新用户编号"""
    r = requests.get(
        f"{BASE_URL}/api/systemcenter/person/newId",
        headers=get_headers(token),
        verify=False, timeout=10
    )
    return r.json()['data']

def create_person(token, person_no: str, name: str, phone: str, group_id: str, group_name: str, remark: str = ""):
    """创建用户"""
    person_data = {
        "personNo": person_no,
        "name": name,
        "mobile": phone,
        "groupId": group_id,
        "groupName": group_name,
        "type": 2,  # 业主
        "gender": "F" if "娣" in name or "女" in remark else "M",
        "relationship": 0,
        "enterTime": datetime.now().strftime("%Y-%m-%dT00:00:00"),
        "status": 1,
        "remark": remark
    }
    r = requests.post(
        f"{BASE_URL}/api/systemcenter/person",
        headers=get_headers(token),
        json=person_data,
        verify=False, timeout=10
    )
    result = r.json()
    if result.get('code') == 200:
        return result['data']['id'], "创建成功"
    return None, f"创建失败: {result}"

def add_credential(token, person_id: str, person_no: str, name: str, plate: str):
    """添加车牌凭证"""
    plate_color = get_plate_color(plate)
    credential_data = {
        "personId": person_id,
        "personNo": person_no,
        "personName": name,
        "credentialNo": plate,
        "credentialType": 163,  # 车牌号码
        "plate": plate,
        "plateColor": plate_color,
        "vehicleType": 0,  # 小型车
        "status": 1
    }
    r = requests.post(
        f"{BASE_URL}/api/systemcenter/credential",
        headers=get_headers(token),
        json=credential_data,
        verify=False, timeout=10
    )
    result = r.json()
    if result.get('code') == 200:
        return result['data']['id'], "凭证添加成功"
    return None, f"添加凭证失败: {result}"

def open_lease_stall(token, person_id: str, person_no: str, name: str, phone: str,
                     group_id: str, group_name: str, credential_id: str, plate: str,
                     seal_id: str, seal_name: str, end_time: str, gender: str = "F"):
    """开通月租车"""
    plate_color = get_plate_color(plate)
    start_time = datetime.now().strftime("%Y-%m-%dT00:00:00")
    
    open_data = {
        "personId": person_id,
        "personNo": person_no,
        "personName": name,
        "mobile": phone,
        "gender": gender,
        "groupId": group_id,
        "groupName": group_name,
        "type": 2,
        "typeName": "业主",
        
        "sealId": seal_id,
        "sealName": seal_name,
        "userType": 1,
        
        "startTime": start_time,
        "endTime": f"{end_time}T23:59:59",
        
        "delayMoney": 0,
        "payTypeID": "XJ",
        
        "carNumber": 1,
        "spaceNumbers": "1",
        "zyPlace": 1,
        "cqPlace": 0,
        "gxPlace": 0,
        "fjdPlace": 0,
        
        "spaceTypeInfoList": [
            {"spaceType": 2, "spaceNumber": 1}
        ],
        
        "credentialNo": plate,
        "credentiallList": [{
            "credentialId": credential_id,
            "credentialNo": plate,
            "credentialType": 163,
            "plateColor": plate_color,
            "vechicleType": "1"
        }]
    }
    
    r = requests.post(
        f"{BASE_URL}/api/parkmanagement/pmsLeaseStall/openLeaseStall",
        headers=get_headers(token),
        json=open_data,
        verify=False, timeout=15
    )
    result = r.json()
    if result.get('code') == 200:
        return True, "月租车开通成功"
    return False, f"开通月租失败: {result}"

def find_lease_by_name(token, name: str):
    """根据姓名查找月租记录"""
    r = requests.get(
        f"{BASE_URL}/api/parkmanagement/pmsLeaseStall/pmsLeaseStallList",
        headers=get_headers(token),
        params={"pageIndex": 0, "pageSize": 10, "personName": name},
        verify=False, timeout=10
    )
    rows = r.json().get('data', {}).get('rows', [])
    return rows[0] if rows else None

def update_plate(token, lease_id: str, person_info: dict, new_plate: str, plate_color: int = None):
    """修改车牌"""
    if plate_color is None:
        plate_color = get_plate_color(new_plate)
    
    body = {
        "id": lease_id,
        "personNo": person_info['personNo'],
        "personName": person_info['personName'],
        "mobile": person_info['mobile'],
        "credentialNo": new_plate,
        "credentiallList": [{
            "credentialId": "",
            "credentialNo": new_plate,
            "credentialType": 163,
            "plateColor": plate_color,
            "vechicleType": "1"
        }],
        "sealId": person_info['sealId'],
        "sealName": person_info['sealName'],
        "userType": 1,
        "startTime": person_info['startTime'],
        "endTime": person_info['endTime'],
        "carNumber": person_info['carNumber'],
        "spaceNumbers": person_info['spaceNumbers'],
        "spaceTypeInfoList": person_info['spaceTypeInfoList'],
        "status": 0
    }
    
    r = requests.put(
        f"{BASE_URL}/api/parkmanagement/pmsLeaseStall/{lease_id}",
        headers=get_headers(token),
        json=body,
        verify=False, timeout=15
    )
    
    result = r.json()
    if result.get('code') == 200:
        sync_update_plate(person_info['personName'], new_plate)
        return True, f"车牌已修改为 {new_plate}"
    return False, f"修改失败: {result}"


# ==================== 本地表格同步 ====================

def sync_update_plate(name: str, new_plate: str):
    """修改车牌 - 同步到本地表格"""
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=COL_NAME).value == name:
            ws.cell(row=row, column=COL_PLATE).value = new_plate
            ws.cell(row=row, column=COL_PLATE).alignment = Alignment(wrap_text=True)
            wb.save(XLSX_PATH)
            print(f"  ✅ 本地表格已更新: {name} -> {new_plate}")
            return True
    
    wb.save(XLSX_PATH)
    print(f"  ⚠️ 本地表格未找到 {name}")
    return False

def sync_add_person(unit: str, name: str, plate: str, phone: str, position: str = "", remark: str = ""):
    """新增人员 - 同步到本地表格"""
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    
    # 检查是否已存在
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=COL_NAME).value == name:
            print(f"  ⚠️ {name} 已存在于本地表格第{row}行")
            wb.close()
            return False
    
    # 找最后有效数据行
    last_data_row = ws.max_row
    while last_data_row > 1 and ws.cell(row=last_data_row, column=COL_NAME).value is None:
        last_data_row -= 1
    
    next_row = last_data_row + 1
    
    # 获取下一个工作证编号
    last_work_id = ws.cell(row=last_data_row, column=COL_WORK_ID).value
    try:
        work_id = int(last_work_id) + 1 if last_work_id else next_row - 1
    except:
        work_id = next_row - 1
    
    ws.cell(row=next_row, column=COL_UNIT, value=unit)
    ws.cell(row=next_row, column=COL_TYPE, value="工作证")
    ws.cell(row=next_row, column=COL_WORK_ID, value=work_id)
    ws.cell(row=next_row, column=COL_NAME, value=name)
    ws.cell(row=next_row, column=COL_POSITION, value=position)
    ws.cell(row=next_row, column=COL_PLATE, value=plate)
    ws.cell(row=next_row, column=COL_PLATE).alignment = Alignment(wrap_text=True)
    ws.cell(row=next_row, column=COL_PHONE, value=phone)
    ws.cell(row=next_row, column=COL_REMARK, value=remark)
    
    wb.save(XLSX_PATH)
    print(f"  ✅ 本地表格已添加: {name} 到第{next_row}行")
    return True


# ==================== 主流程 ====================

def modify_plate(name: str, new_plate: str):
    """修改车牌（完整流程：API + 本地表格）"""
    print(f"\n{'='*50}")
    print(f"修改车牌: {name} -> {new_plate}")
    print(f"{'='*50}")
    
    token = login()
    print("[1/3] ✅ 登录成功")
    
    lease = find_lease_by_name(token, name)
    if not lease:
        print(f"[2/3] ❌ 未找到 {name} 的月租记录")
        return False
    print(f"[2/3] ✅ 找到月租记录")
    
    success, msg = update_plate(token, lease['id'], lease, new_plate)
    print(f"[3/3] {'✅' if success else '❌'} {msg}")
    
    return success

def add_person_full(unit: str, name: str, plate: str, phone: str, 
                    seal_name: str = "地面月卡", end_time: str = "2030-12-30", 
                    remark: str = ""):
    """新增人员（完整流程：API + 本地表格）"""
    print(f"\n{'='*50}")
    print(f"新增人员: {name}")
    print(f"  单位: {unit}")
    print(f"  车牌: {plate}")
    print(f"  电话: {phone}")
    print(f"  套餐: {seal_name}")
    print(f"  有效期: {end_time}")
    print(f"{'='*50}")
    
    # 1. 登录
    token = login()
    print("[1/6] ✅ 登录成功")
    
    # 2. 查找组织
    group_id, group_name = find_group_id(token, unit)
    if not group_id:
        print(f"[2/6] ❌ 未找到组织: {unit}")
        return False
    print(f"[2/6] ✅ 组织: {group_name}")
    
    # 3. 获取新用户编号
    person_no = get_new_person_no(token)
    print(f"[3/6] ✅ 用户编号: {person_no}")
    
    # 4. 创建用户
    gender = "F" if "娣" in name or "女" in remark else "M"
    person_id, msg = create_person(token, person_no, name, phone, group_id, group_name, remark)
    if not person_id:
        print(f"[4/6] ❌ {msg}")
        return False
    print(f"[4/6] ✅ 用户创建成功")
    
    # 5. 添加车牌凭证
    credential_id, msg = add_credential(token, person_id, person_no, name, plate)
    if not credential_id:
        print(f"[5/6] ❌ {msg}")
        return False
    print(f"[5/6] ✅ 车牌凭证添加成功")
    
    # 6. 开通月租车
    seal_id = SEAL_MAP.get(seal_name)
    if not seal_id:
        print(f"[6/6] ⚠️ 未知套餐: {seal_name}，尝试使用名称")
        seal_id = seal_name
    
    success, msg = open_lease_stall(token, person_id, person_no, name, phone,
                                    group_id, group_name, credential_id, plate,
                                    seal_id, seal_name, end_time, gender)
    if not success:
        print(f"[6/6] ❌ {msg}")
        return False
    print(f"[6/6] ✅ 月租车开通成功")
    
    # 7. 同步本地表格
    sync_add_person(unit, name, plate, phone, "", remark)
    
    print(f"\n{'='*50}")
    print(f"✅ 全部完成！")
    print(f"{'='*50}")
    return True


# ==================== 入口 ====================

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    
    cmd = sys.argv[1]
    
    if cmd == "modify" and len(sys.argv) >= 4:
        name = sys.argv[2]
        new_plate = sys.argv[3]
        modify_plate(name, new_plate)
        
    elif cmd == "add" and len(sys.argv) >= 7:
        unit = sys.argv[2]
        name = sys.argv[3]
        plate = sys.argv[4]
        phone = sys.argv[5]
        seal_name = sys.argv[6]
        end_time = sys.argv[7] if len(sys.argv) > 7 else "2030-12-30"
        remark = sys.argv[8] if len(sys.argv) > 8 else ""
        add_person_full(unit, name, plate, phone, seal_name, end_time, remark)
        
    else:
        print("参数错误，请查看使用方法：")
        print(__doc__)
        sys.exit(1)
