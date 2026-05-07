#!/usr/bin/env python3
"""
停车场管理系统 - PyQt5桌面应用
功能：修改车牌、新增人员、黑白名单管理、本地Excel同步
"""

import hashlib
import json
import sys
import urllib3
from datetime import datetime

import requests
import openpyxl
from openpyxl.styles import Alignment
from PyQt5.QtCore import (
    QDate, QMutex, QObject, QRunnable, QSettings, Qt, QThreadPool,
    pyqtSignal
)
from PyQt5.QtWidgets import (
    QApplication, QCheckBox, QComboBox, QDateEdit, QDialog,
    QFileDialog, QFormLayout, QGroupBox, QHBoxLayout, QHeaderView,
    QLabel, QLineEdit, QMainWindow, QMessageBox, QProgressBar,
    QPushButton, QSpinBox, QStatusBar, QTabWidget, QTableWidget,
    QTableWidgetItem, QTextEdit, QVBoxLayout, QWidget
)

urllib3.disable_warnings()

# ==================== 常量配置 ====================

SEAL_MAP = {
    "地面月卡": "p220447822150",
    "地面临时车": "p220447822154",
    "地库临时车": "p220447822155",
    "地库月卡": "p220447822159",
}

LIST_TYPE_MAP = {
    "黑名单": 1,
    "灰名单": 2,
    "白名单": 3,
}

COL_UNIT = 1
COL_TYPE = 2
COL_WORK_ID = 3
COL_NAME = 4
COL_POSITION = 5
COL_PLATE = 6
COL_PHONE = 7
COL_REMARK = 8


def get_plate_color(plate: str) -> int:
    if not plate or len(plate) < 2:
        return 3
    if plate.startswith(("粤B", "粤D", "粤F")) and len(plate) == 8:
        return 5
    return 3


# ==================== 后台线程 ====================

class WorkerSignals(QObject):
    success = pyqtSignal(object)
    error = pyqtSignal(str)
    log = pyqtSignal(str)


class Worker(QRunnable):
    def __init__(self, func, *args, **kwargs):
        super().__init__()
        self.func = func
        self.args = args
        self.kwargs = kwargs
        self.signals = WorkerSignals()
        self.setAutoDelete(True)

    def run(self):
        try:
            result = self.func(*self.args, **self.kwargs)
            self.signals.success.emit(result)
        except Exception as e:
            self.signals.error.emit(str(e))


# ==================== 停车场API ====================

class ParkingAPI:
    def __init__(self, base_url: str, account: str, password: str):
        self.base_url = base_url.rstrip("/")
        self.account = account
        self.password = password
        self.token = None
        self.groups = []

    def _headers(self):
        return {"Authorization": f"Bearer {self.token}", "Content-Type": "application/json"}

    def login(self) -> dict:
        pw_md5 = hashlib.md5(self.password.encode()).hexdigest()
        r = requests.post(
            f"{self.base_url}/api/systemcenter/auth/login",
            json={"account": self.account, "password": pw_md5},
            verify=False, timeout=10
        )
        data = r.json()
        if data.get("code") == 200:
            self.token = data["data"]["token"]
            return {"ok": True, "msg": "登录成功"}
        return {"ok": False, "msg": f"登录失败: {data}"}

    def fetch_groups(self) -> list:
        all_rows = []
        page = 0
        while True:
            r = requests.get(
                f"{self.base_url}/api/systemcenter/group",
                headers=self._headers(),
                params={"pageIndex": page, "pageSize": 100},
                verify=False, timeout=10
            )
            data = r.json().get("data", {})
            rows = data.get("rows", [])
            all_rows.extend(rows)
            total = data.get("total", 0)
            if len(all_rows) >= total or not rows:
                break
            page += 1
        self.groups = all_rows
        return all_rows

    def find_group_id(self, hint: str):
        hint = hint.strip()
        # 精确匹配
        for g in self.groups:
            if g.get("name", "") == hint:
                return g["id"], g["name"]
        # 用户输入包含在组织名中（如 "三联村委" in "三联村民委员会"）
        for g in self.groups:
            if hint in g.get("name", ""):
                return g["id"], g["name"]
        # 组织名包含在用户输入中（如用户输入 "三联村民委员会"，组织名为 "三联村委"）
        for g in self.groups:
            if g.get("name", "") and g["name"] in hint:
                return g["id"], g["name"]
        return None, None

    def get_new_person_no(self) -> str:
        r = requests.get(
            f"{self.base_url}/api/systemcenter/person/newId",
            headers=self._headers(),
            verify=False, timeout=10
        )
        return r.json()["data"]

    def create_person(self, person_no, name, phone, group_id, group_name, remark=""):
        gender = "F" if "娣" in name or "女" in remark else "M"
        data = {
            "personNo": person_no, "name": name, "mobile": phone,
            "groupId": group_id, "groupName": group_name,
            "type": 2, "gender": gender, "relationship": 0,
            "enterTime": datetime.now().strftime("%Y-%m-%dT00:00:00"),
            "status": 1, "remark": remark
        }
        r = requests.post(
            f"{self.base_url}/api/systemcenter/person",
            headers=self._headers(), json=data, verify=False, timeout=10
        )
        res = r.json()
        if res.get("code") == 200:
            return res["data"]["id"], "创建成功"
        return None, f"创建失败: {res}"

    def add_credential(self, person_id, person_no, name, plate):
        plate_color = get_plate_color(plate)
        data = {
            "personId": person_id, "personNo": person_no, "personName": name,
            "credentialNo": plate, "credentialType": 163,
            "plate": plate, "plateColor": plate_color,
            "vehicleType": 0, "status": 1
        }
        r = requests.post(
            f"{self.base_url}/api/systemcenter/credential",
            headers=self._headers(), json=data, verify=False, timeout=10
        )
        res = r.json()
        if res.get("code") == 200:
            return res["data"]["id"], "凭证添加成功"
        return None, f"添加凭证失败: {res}"

    def open_lease_stall(self, person_id, person_no, name, phone,
                         group_id, group_name, credential_id, plate,
                         seal_id, seal_name, end_time, gender="M"):
        plate_color = get_plate_color(plate)
        start_time = datetime.now().strftime("%Y-%m-%dT00:00:00")
        data = {
            "personId": person_id, "personNo": person_no, "personName": name,
            "mobile": phone, "gender": gender,
            "groupId": group_id, "groupName": group_name,
            "type": 2, "typeName": "业主",
            "sealId": seal_id, "sealName": seal_name, "userType": 1,
            "startTime": start_time, "endTime": f"{end_time}T23:59:59",
            "delayMoney": 0, "payTypeID": "XJ",
            "carNumber": 1, "spaceNumbers": "1",
            "zyPlace": 1, "cqPlace": 0, "gxPlace": 0, "fjdPlace": 0,
            "spaceTypeInfoList": [{"spaceType": 2, "spaceNumber": 1}],
            "credentialNo": plate,
            "credentiallList": [{
                "credentialId": credential_id, "credentialNo": plate,
                "credentialType": 163, "plateColor": plate_color, "vechicleType": "1"
            }]
        }
        r = requests.post(
            f"{self.base_url}/api/parkmanagement/pmsLeaseStall/openLeaseStall",
            headers=self._headers(), json=data, verify=False, timeout=15
        )
        res = r.json()
        if res.get("code") == 200:
            return True, "月租车开通成功"
        return False, f"开通月租失败: {res}"

    def find_lease_by_name(self, name: str):
        r = requests.get(
            f"{self.base_url}/api/parkmanagement/pmsLeaseStall/pmsLeaseStallList",
            headers=self._headers(),
            params={"pageIndex": 0, "pageSize": 10, "personName": name},
            verify=False, timeout=10
        )
        rows = r.json().get("data", {}).get("rows", [])
        return rows[0] if rows else None

    def update_plate(self, lease_id, person_info, new_plate):
        plate_color = get_plate_color(new_plate)
        body = {
            "id": lease_id,
            "personNo": person_info["personNo"],
            "personName": person_info["personName"],
            "mobile": person_info["mobile"],
            "credentialNo": new_plate,
            "credentiallList": [{
                "credentialId": "", "credentialNo": new_plate,
                "credentialType": 163, "plateColor": plate_color, "vechicleType": "1"
            }],
            "sealId": person_info["sealId"],
            "sealName": person_info["sealName"],
            "userType": 1,
            "startTime": person_info["startTime"],
            "endTime": person_info["endTime"],
            "carNumber": person_info["carNumber"],
            "spaceNumbers": person_info["spaceNumbers"],
            "spaceTypeInfoList": person_info["spaceTypeInfoList"],
            "status": 0
        }
        r = requests.put(
            f"{self.base_url}/api/parkmanagement/pmsLeaseStall/{lease_id}",
            headers=self._headers(), json=body, verify=False, timeout=15
        )
        res = r.json()
        if res.get("code") == 200:
            return True, f"车牌已修改为 {new_plate}"
        return False, f"修改失败: {res}"

    # ---- 黑白名单 ----

    def add_bw_record(self, plate, list_type, start_date, end_date, remark=""):
        data = {
            "credentialNo": plate,
            "credentialType": 163,
            "plate": plate,
            "plateColor": get_plate_color(plate),
            "vehicleType": 0,
            "listType": list_type,
            "startTime": f"{start_date}T00:00:00",
            "endTime": f"{end_date}T23:59:59",
            "status": 1,
            "remark": remark
        }
        r = requests.post(
            f"{self.base_url}/api/parkmanagement/pmsBlackWhiteList",
            headers=self._headers(), json=data, verify=False, timeout=10
        )
        res = r.json()
        if res.get("code") == 200:
            return True, "添加成功"
        return False, f"添加失败: {res}"

    def query_bw_list(self, plate="", list_type=None, page_size=50):
        params = {"pageIndex": 0, "pageSize": page_size}
        if plate:
            params["credentialNo"] = plate
        if list_type is not None:
            params["listType"] = list_type
        r = requests.get(
            f"{self.base_url}/api/parkmanagement/pmsBlackWhiteList",
            headers=self._headers(), params=params, verify=False, timeout=10
        )
        res = r.json()
        if res.get("code") == 200:
            return res.get("data", {}).get("rows", [])
        return []

    def delete_bw_record(self, record_id):
        r = requests.delete(
            f"{self.base_url}/api/parkmanagement/pmsBlackWhiteList/{record_id}",
            headers=self._headers(), verify=False, timeout=10
        )
        res = r.json()
        if res.get("code") == 200:
            return True, "删除成功"
        return False, f"删除失败: {res}"

    # ---- 凭证管理 & 用户删除 ----

    def query_credential(self, keyword, page_size=20):
        params = {"pageIndex": 0, "pageSize": page_size}
        if keyword:
            params["credentialNo"] = keyword
        r = requests.get(
            f"{self.base_url}/api/systemcenter/credential",
            headers=self._headers(), params=params, verify=False, timeout=10
        )
        res = r.json()
        if res.get("code") == 200:
            return res.get("data", {}).get("rows", [])
        return []

    def query_credential_by_name(self, name, page_size=20):
        params = {"pageIndex": 0, "pageSize": page_size}
        if name:
            params["personName"] = name
        r = requests.get(
            f"{self.base_url}/api/systemcenter/credential",
            headers=self._headers(), params=params, verify=False, timeout=10
        )
        res = r.json()
        if res.get("code") == 200:
            return res.get("data", {}).get("rows", [])
        return []

    def cancel_credential(self, credential_ids):
        if isinstance(credential_ids, str):
            credential_ids = [credential_ids]
        data = {"ids": credential_ids, "status": 4}
        r = requests.patch(
            f"{self.base_url}/api/systemcenter/credential",
            headers=self._headers(), json=data, verify=False, timeout=10
        )
        res = r.json()
        if res.get("code") == 200:
            return True, "凭证注销成功"
        return False, f"凭证注销失败: {res}"

    def query_person(self, name, page_size=20):
        params = {"pageIndex": 0, "pageSize": page_size}
        if name:
            params["name"] = name
        r = requests.get(
            f"{self.base_url}/api/systemcenter/person",
            headers=self._headers(), params=params, verify=False, timeout=10
        )
        res = r.json()
        if res.get("code") == 200:
            return res.get("data", {}).get("rows", [])
        return []

    def batch_migrate_out(self, person_ids):
        if isinstance(person_ids, str):
            person_ids = [person_ids]
        data = {"ids": person_ids, "isServicesCancelled": True}
        r = requests.patch(
            f"{self.base_url}/api/systemcenter/person/moveOut",
            headers=self._headers(), json=data, verify=False, timeout=15
        )
        res = r.json()
        if res.get("code") == 200:
            return True, "批量迁出成功（用户信息及关联凭证已删除）"
        return False, f"批量迁出失败: {res}"


# ==================== Excel同步 ====================

def sync_update_plate(xlsx_path, name, new_plate):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=COL_NAME).value == name:
            ws.cell(row=row, column=COL_PLATE).value = new_plate
            ws.cell(row=row, column=COL_PLATE).alignment = Alignment(wrap_text=True)
            wb.save(xlsx_path)
            return True, f"本地表格已更新: {name} -> {new_plate}"
    wb.save(xlsx_path)
    return False, f"本地表格未找到 {name}"


def sync_add_person(xlsx_path, unit, name, plate, phone, remark=""):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=COL_NAME).value == name:
            wb.close()
            return False, f"{name} 已存在于本地表格第{row}行"
    last_row = ws.max_row
    while last_row > 1 and ws.cell(row=last_row, column=COL_NAME).value is None:
        last_row -= 1
    next_row = last_row + 1
    last_work_id = ws.cell(row=last_row, column=COL_WORK_ID).value
    try:
        work_id = int(last_work_id) + 1 if last_work_id else next_row - 1
    except (ValueError, TypeError):
        work_id = next_row - 1
    ws.cell(row=next_row, column=COL_UNIT, value=unit)
    ws.cell(row=next_row, column=COL_TYPE, value="工作证")
    ws.cell(row=next_row, column=COL_WORK_ID, value=work_id)
    ws.cell(row=next_row, column=COL_NAME, value=name)
    ws.cell(row=next_row, column=COL_POSITION, value="")
    ws.cell(row=next_row, column=COL_PLATE, value=plate)
    ws.cell(row=next_row, column=COL_PLATE).alignment = Alignment(wrap_text=True)
    ws.cell(row=next_row, column=COL_PHONE, value=phone)
    ws.cell(row=next_row, column=COL_REMARK, value=remark)
    wb.save(xlsx_path)
    return True, f"本地表格已添加: {name} 到第{next_row}行"


def sync_delete_person(xlsx_path, name):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=COL_NAME).value == name:
            ws.delete_rows(row)
            wb.save(xlsx_path)
            return True, f"本地表格已删除: {name} (第{row}行)"
    wb.save(xlsx_path)
    return False, f"本地表格未找到 {name}"


# ==================== 登录对话框 ====================

class LoginDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("停车场管理系统 - 登录")
        self.setFixedSize(420, 300)
        self._setup_ui()
        self._load_settings()

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        form = QFormLayout()
        self.server_edit = QLineEdit("https://10.0.12.1:9091")
        self.account_edit = QLineEdit("9999")
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.Password)

        xlsx_row = QHBoxLayout()
        self.xlsx_edit = QLineEdit()
        self.xlsx_edit.setPlaceholderText("选择Excel汇总表路径...")
        browse_btn = QPushButton("浏览")
        browse_btn.setFixedWidth(60)
        browse_btn.clicked.connect(self._browse_xlsx)
        xlsx_row.addWidget(self.xlsx_edit)
        xlsx_row.addWidget(browse_btn)

        self.remember_check = QCheckBox("记住凭据")

        form.addRow("服务器地址:", self.server_edit)
        form.addRow("账  号:", self.account_edit)
        form.addRow("密  码:", self.password_edit)
        form.addRow("Excel路径:", xlsx_row)
        form.addRow(self.remember_check)
        layout.addLayout(form)

        self.login_btn = QPushButton("登  录")
        self.login_btn.setFixedHeight(36)
        self.login_btn.clicked.connect(self._on_login)
        layout.addWidget(self.login_btn)

        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: red;")
        layout.addWidget(self.status_label)

    def _load_settings(self):
        s = QSettings("ParkingSystem", "ParkingGUI")
        self.server_edit.setText(s.value("server", "https://10.0.12.1:9091"))
        self.account_edit.setText(s.value("account", "9999"))
        self.xlsx_edit.setText(s.value("xlsx_path", ""))
        if s.value("remember", "false") == "true":
            self.remember_check.setChecked(True)
            self.password_edit.setText(s.value("password", ""))

    def _save_settings(self):
        s = QSettings("ParkingSystem", "ParkingGUI")
        s.setValue("server", self.server_edit.text().strip())
        s.setValue("account", self.account_edit.text().strip())
        s.setValue("xlsx_path", self.xlsx_edit.text().strip())
        s.setValue("remember", "true" if self.remember_check.isChecked() else "false")
        if self.remember_check.isChecked():
            s.setValue("password", self.password_edit.text().strip())

    def _browse_xlsx(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)"
        )
        if path:
            self.xlsx_edit.setText(path)

    def _on_login(self):
        server = self.server_edit.text().strip()
        account = self.account_edit.text().strip()
        password = self.password_edit.text().strip()
        xlsx = self.xlsx_edit.text().strip()

        if not all([server, account, password]):
            self.status_label.setText("请填写完整登录信息")
            return

        self.login_btn.setEnabled(False)
        self.status_label.setText("正在登录...")
        QApplication.processEvents()

        api = ParkingAPI(server, account, password)
        result = api.login()

        if result["ok"]:
            self._save_settings()
            self.api = api
            self.xlsx_path = xlsx
            self.accept()
        else:
            self.status_label.setText(result["msg"])
            self.login_btn.setEnabled(True)


# ==================== 主窗口 ====================

class MainWindow(QMainWindow):
    def __init__(self, api: ParkingAPI, xlsx_path: str):
        super().__init__()
        self.api = api
        self.xlsx_path = xlsx_path
        self.thread_pool = QThreadPool()
        self.setWindowTitle("停车场管理系统")
        self.resize(700, 580)
        self._setup_ui()
        self._fetch_groups()

    def _setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)

        self.tabs = QTabWidget()
        self.tabs.addTab(ModifyPlateTab(self), "修改车牌")
        self.tabs.addTab(AddPersonTab(self), "新增人员")
        self.tabs.addTab(DeletePersonTab(self), "删除人员")
        self.tabs.addTab(BlackWhiteListTab(self), "黑白名单")
        layout.addWidget(self.tabs)

        log_group = QGroupBox("操作日志")
        log_layout = QVBoxLayout(log_group)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(140)
        log_layout.addWidget(self.log_text)
        layout.addWidget(log_group)

        self.statusBar().showMessage(f"已登录 | 服务器: {self.api.base_url}")

    def log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{ts}] {msg}")

    def run_worker(self, func, on_success, on_error=None):
        worker = Worker(func)
        worker.signals.success.connect(on_success)
        worker.signals.error.connect(on_error or (lambda e: self.log(f"错误: {e}")))
        self.thread_pool.start(worker)

    def _fetch_groups(self):
        self.log("获取组织列表...")

        def do_fetch():
            return self.api.fetch_groups()

        def on_ok(groups):
            names = [g["name"] for g in groups]
            self.log(f"获取到 {len(names)} 个组织")
            add_tab = self.tabs.widget(1)
            if isinstance(add_tab, AddPersonTab):
                add_tab.populate_groups(names)

        self.run_worker(do_fetch, on_ok)


# ==================== 修改车牌Tab ====================

class ModifyPlateTab(QWidget):
    def __init__(self, main: MainWindow):
        super().__init__()
        self.main = main
        self.current_lease = None
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        form = QFormLayout()
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("输入姓名查询")
        self.new_plate_edit = QLineEdit()
        self.new_plate_edit.setPlaceholderText("例: 粤S22222")

        search_btn = QPushButton("查询")
        search_btn.setFixedWidth(60)
        search_btn.clicked.connect(self._on_search)

        name_row = QHBoxLayout()
        name_row.addWidget(self.name_edit)
        name_row.addWidget(search_btn)
        form.addRow("姓  名:", name_row)
        form.addRow("新车牌:", self.new_plate_edit)

        self.info_label = QLabel("")
        self.info_label.setStyleSheet("color: #1565C0; font-weight: bold;")
        form.addRow("当前信息:", self.info_label)

        self.sync_check = QCheckBox("同步更新本地Excel")
        self.sync_check.setChecked(True)
        form.addRow(self.sync_check)

        layout.addLayout(form)

        self.modify_btn = QPushButton("确认修改")
        self.modify_btn.setFixedHeight(36)
        self.modify_btn.setEnabled(False)
        self.modify_btn.clicked.connect(self._on_modify)
        layout.addWidget(self.modify_btn)
        layout.addStretch()

    def _on_search(self):
        name = self.name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "提示", "请输入姓名")
            return
        self.modify_btn.setEnabled(False)
        self.info_label.setText("查询中...")
        self.main.log(f"查询月租记录: {name}")

        def do_search():
            return self.main.api.find_lease_by_name(name)

        def on_ok(lease):
            self.current_lease = lease
            if lease:
                plate = lease.get("credentialNo", "未知")
                seal = lease.get("sealName", "")
                end = lease.get("endTime", "")[:10]
                self.info_label.setText(f"车牌: {plate} | 套餐: {seal} | 到期: {end}")
                self.modify_btn.setEnabled(True)
                self.main.log(f"找到记录: {plate}")
            else:
                self.info_label.setText("未找到该人员的月租记录")
                self.main.log("未找到记录")

        self.main.run_worker(do_search, on_ok)

    def _on_modify(self):
        new_plate = self.new_plate_edit.text().strip()
        if not new_plate:
            QMessageBox.warning(self, "提示", "请输入新车牌")
            return
        if not self.current_lease:
            return

        self.modify_btn.setEnabled(False)
        name = self.name_edit.text().strip()
        self.main.log(f"修改车牌: {name} -> {new_plate}")

        lease = self.current_lease

        def do_modify():
            return self.main.api.update_plate(lease["id"], lease, new_plate)

        def on_ok(result):
            ok, msg = result
            self.main.log(msg)
            if ok:
                # 同步Excel
                if self.sync_check.isChecked() and self.main.xlsx_path:
                    try:
                        sync_ok, sync_msg = sync_update_plate(
                            self.main.xlsx_path, name, new_plate
                        )
                        self.main.log(sync_msg)
                    except Exception as e:
                        self.main.log(f"Excel同步失败: {e}")
                QMessageBox.information(self, "成功", f"车牌修改成功\n{msg}")
                self.info_label.setText(f"车牌: {new_plate}")
                self.current_lease["credentialNo"] = new_plate
            else:
                QMessageBox.warning(self, "失败", msg)
            self.modify_btn.setEnabled(True)

        self.main.run_worker(do_modify, on_ok)


# ==================== 新增人员Tab ====================

class AddPersonTab(QWidget):
    def __init__(self, main: MainWindow):
        super().__init__()
        self.main = main
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        form = QFormLayout()
        self.unit_combo = QComboBox()
        self.unit_combo.setEditable(True)
        self.unit_combo.setPlaceholderText("选择或输入单位名称")

        self.name_edit = QLineEdit()
        self.plate_edit = QLineEdit()
        self.plate_edit.setPlaceholderText("例: 粤BF78K9")
        self.phone_edit = QLineEdit()
        self.phone_edit.setPlaceholderText("手机号码")

        self.seal_combo = QComboBox()
        self.seal_combo.addItems(SEAL_MAP.keys())

        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.end_date_edit.setDate(QDate(2030, 12, 30))

        self.remark_edit = QLineEdit()
        self.remark_edit.setPlaceholderText("选填")

        self.sync_check = QCheckBox("同步更新本地Excel")
        self.sync_check.setChecked(True)

        form.addRow("单  位:", self.unit_combo)
        form.addRow("姓  名:", self.name_edit)
        form.addRow("车  牌:", self.plate_edit)
        form.addRow("电  话:", self.phone_edit)
        form.addRow("套  餐:", self.seal_combo)
        form.addRow("有效期至:", self.end_date_edit)
        form.addRow("备  注:", self.remark_edit)
        form.addRow(self.sync_check)
        layout.addLayout(form)

        self.progress = QProgressBar()
        self.progress.setMaximum(7)
        self.progress.setValue(0)
        layout.addWidget(self.progress)

        self.add_btn = QPushButton("新增人员")
        self.add_btn.setFixedHeight(36)
        self.add_btn.clicked.connect(self._on_add)
        layout.addWidget(self.add_btn)
        layout.addStretch()

    def populate_groups(self, names):
        self.unit_combo.clear()
        self.unit_combo.addItems(names)

    def _on_add(self):
        unit = self.unit_combo.currentText().strip()
        name = self.name_edit.text().strip()
        plate = self.plate_edit.text().strip()
        phone = self.phone_edit.text().strip()
        seal_name = self.seal_combo.currentText()
        end_date = self.end_date_edit.date().toString("yyyy-MM-dd")
        remark = self.remark_edit.text().strip()

        if not all([unit, name, plate, phone]):
            QMessageBox.warning(self, "提示", "请填写单位、姓名、车牌、电话")
            return

        self.add_btn.setEnabled(False)
        self.progress.setValue(0)
        self.main.log(f"新增人员: {name} | {unit} | {plate}")

        def do_add():
            api = self.main.api
            xlsx = self.main.xlsx_path
            logs = []

            # 1. 登录
            logs.append(("[1/7] 登录中..."))
            api.login()
            logs.append(("[1/7] 登录成功"))

            # 2. 查找组织
            logs.append(("[2/7] 查找组织..."))
            group_id, group_name = api.find_group_id(unit)
            if not group_id:
                return False, "未找到组织: " + unit, logs
            logs.append(f"[2/7] 组织: {group_name}")

            # 3. 获取新编号
            logs.append(("[3/7] 获取编号..."))
            person_no = api.get_new_person_no()
            logs.append(f"[3/7] 编号: {person_no}")

            # 4. 创建用户
            logs.append(("[4/7] 创建用户..."))
            gender = "F" if "娣" in name or "女" in remark else "M"
            person_id, msg = api.create_person(
                person_no, name, phone, group_id, group_name, remark
            )
            if not person_id:
                return False, msg, logs
            logs.append(f"[4/7] 用户创建成功")

            # 5. 添加凭证
            logs.append(("[5/7] 添加凭证..."))
            credential_id, msg = api.add_credential(person_id, person_no, name, plate)
            if not credential_id:
                return False, msg, logs
            logs.append("[5/7] 凭证添加成功")

            # 6. 开通月租
            logs.append(("[6/7] 开通月租..."))
            seal_id = SEAL_MAP.get(seal_name, seal_name)
            ok, msg = api.open_lease_stall(
                person_id, person_no, name, phone, group_id, group_name,
                credential_id, plate, seal_id, seal_name, end_date, gender
            )
            if not ok:
                return False, msg, logs
            logs.append("[6/7] 月租车开通成功")

            # 7. 同步Excel
            logs.append(("[7/7] 同步Excel..."))
            if xlsx:
                try:
                    sync_ok, sync_msg = sync_add_person(xlsx, unit, name, plate, phone, remark)
                    logs.append(f"[7/7] {sync_msg}")
                except Exception as e:
                    logs.append(f"[7/7] Excel同步失败: {e}")
            else:
                logs.append("[7/7] 未配置Excel路径，跳过同步")

            return True, "新增人员完成", logs

        def on_ok(result):
            ok, msg, logs = result
            for l in logs:
                self.main.log(l)
            if ok:
                self.progress.setValue(7)
                QMessageBox.information(self, "成功", msg)
                self.name_edit.clear()
                self.plate_edit.clear()
                self.phone_edit.clear()
                self.remark_edit.clear()
            else:
                QMessageBox.warning(self, "失败", msg)
            self.add_btn.setEnabled(True)

        def on_step_progress():
            val = self.progress.value() + 1
            self.progress.setValue(min(val, 7))

        self.main.run_worker(do_add, on_ok)


# ==================== 删除人员Tab ====================

class DeletePersonTab(QWidget):
    def __init__(self, main: MainWindow):
        super().__init__()
        self.main = main
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        # 操作类型
        type_group = QGroupBox("删除方式")
        type_layout = QVBoxLayout(type_group)
        self.type_combo = QComboBox()
        self.type_combo.addItems([
            "按姓名删除（用户管理-批量迁出，同时删除用户和车牌）",
            "按车牌注销并删除（先注销车牌，再删除用户）"
        ])
        self.type_combo.currentIndexChanged.connect(self._on_type_changed)
        type_layout.addWidget(self.type_combo)
        layout.addWidget(type_group)

        # 搜索区
        search_group = QGroupBox("搜索")
        search_layout = QFormLayout(search_group)
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("输入姓名")
        self.plate_edit = QLineEdit()
        self.plate_edit.setPlaceholderText("输入车牌号，例: 粤S12345")

        name_search_btn = QPushButton("搜索")
        name_search_btn.setFixedWidth(60)
        name_search_btn.clicked.connect(self._on_search_name)
        name_row = QHBoxLayout()
        name_row.addWidget(self.name_edit)
        name_row.addWidget(name_search_btn)
        search_layout.addRow("姓  名:", name_row)

        plate_search_btn = QPushButton("搜索")
        plate_search_btn.setFixedWidth(60)
        plate_search_btn.clicked.connect(self._on_search_plate)
        plate_row = QHBoxLayout()
        plate_row.addWidget(self.plate_edit)
        plate_row.addWidget(plate_search_btn)
        search_layout.addRow("车  牌:", plate_row)

        layout.addWidget(search_group)

        # 搜索结果
        self.result_label = QLabel("")
        self.result_label.setStyleSheet("color: #1565C0; font-weight: bold;")
        layout.addWidget(self.result_label)

        self.credential_table = QTableWidget(0, 4)
        self.credential_table.setHorizontalHeaderLabels(["凭证ID", "车牌号", "姓名", "状态"])
        self.credential_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.credential_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.credential_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.credential_table.setMaximumHeight(120)
        layout.addWidget(self.credential_table)

        self.person_table = QTableWidget(0, 4)
        self.person_table.setHorizontalHeaderLabels(["用户ID", "姓名", "手机", "组织"])
        self.person_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.person_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.person_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.person_table.setMaximumHeight(120)
        layout.addWidget(self.person_table)

        # 同步选项
        self.sync_check = QCheckBox("同步删除本地Excel记录")
        self.sync_check.setChecked(True)
        layout.addWidget(self.sync_check)

        # 执行按钮
        self.delete_btn = QPushButton("执行删除")
        self.delete_btn.setFixedHeight(36)
        self.delete_btn.setEnabled(False)
        self.delete_btn.setStyleSheet("QPushButton { background-color: #D32F2F; color: white; font-weight: bold; }")
        self.delete_btn.clicked.connect(self._on_delete)
        layout.addWidget(self.delete_btn)

        # 内部状态
        self._found_credentials = []
        self._found_persons = []
        self._delete_mode = "name"  # "name" or "plate"

        layout.addStretch()

    def _on_type_changed(self, index):
        self._delete_mode = "name" if index == 0 else "plate"
        self.name_edit.setEnabled(self._delete_mode == "name")
        self.plate_edit.setEnabled(self._delete_mode == "plate")
        self._clear_results()

    def _clear_results(self):
        self.credential_table.setRowCount(0)
        self.person_table.setRowCount(0)
        self.result_label.setText("")
        self.delete_btn.setEnabled(False)
        self._found_credentials = []
        self._found_persons = []

    def _on_search_name(self):
        name = self.name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "提示", "请输入姓名")
            return
        self._clear_results()
        self.result_label.setText("搜索中...")
        self.main.log(f"搜索用户: {name}")

        def do_search():
            persons = self.main.api.query_person(name)
            return persons

        def on_ok(persons):
            self._found_persons = persons
            self.person_table.setRowCount(len(persons))
            for i, p in enumerate(persons):
                self.person_table.setItem(i, 0, QTableWidgetItem(str(p.get("id", ""))))
                self.person_table.setItem(i, 1, QTableWidgetItem(p.get("name", "")))
                self.person_table.setItem(i, 2, QTableWidgetItem(p.get("mobile", "")))
                self.person_table.setItem(i, 3, QTableWidgetItem(p.get("groupName", "")))
            if persons:
                self.result_label.setText(f"找到 {len(persons)} 个用户")
                self.delete_btn.setEnabled(True)
                self.main.log(f"找到 {len(persons)} 个用户")
            else:
                self.result_label.setText("未找到该用户")
                self.main.log("未找到用户")

        self.main.run_worker(do_search, on_ok)

    def _on_search_plate(self):
        plate = self.plate_edit.text().strip()
        if not plate:
            QMessageBox.warning(self, "提示", "请输入车牌号")
            return
        self._clear_results()
        self.result_label.setText("搜索中...")
        self.main.log(f"搜索凭证: {plate}")

        def do_search():
            creds = self.main.api.query_credential(plate)
            return creds

        def on_ok(creds):
            self._found_credentials = creds
            self.credential_table.setRowCount(len(creds))
            for i, c in enumerate(creds):
                self.credential_table.setItem(i, 0, QTableWidgetItem(str(c.get("id", ""))))
                self.credential_table.setItem(i, 1, QTableWidgetItem(c.get("credentialNo", "")))
                self.credential_table.setItem(i, 2, QTableWidgetItem(c.get("personName", "")))
                status = "正常" if c.get("status") == 1 else "已注销"
                self.credential_table.setItem(i, 3, QTableWidgetItem(status))
            if creds:
                person_name = creds[0].get("personName", "")
                self.name_edit.setText(person_name)
                self.result_label.setText(
                    f"找到 {len(creds)} 条凭证 | 车主: {person_name}"
                )
                self.delete_btn.setEnabled(True)
                self.main.log(f"找到凭证: {person_name}")
            else:
                self.result_label.setText("未找到该车牌凭证")
                self.main.log("未找到凭证")

        self.main.run_worker(do_search, on_ok)

    def _on_delete(self):
        if self._delete_mode == "name":
            self._delete_by_name()
        else:
            self._delete_by_plate()

    def _delete_by_name(self):
        if not self._found_persons:
            return
        names = [p.get("name", "") for p in self._found_persons]
        person_ids = [p.get("id") for p in self._found_persons]
        name = self.name_edit.text().strip()

        confirm = QMessageBox.question(
            self, "确认删除",
            f"将批量迁出以下用户（同时删除用户信息和关联车牌）:\n"
            f"{', '.join(names)}\n\n"
            f"此操作不可撤销，确定继续？",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return

        self.delete_btn.setEnabled(False)
        self.main.log(f"批量迁出用户: {', '.join(names)}")

        def do_delete():
            api = self.main.api
            logs = []
            ok, msg = api.batch_migrate_out(person_ids)
            logs.append(msg)
            return ok, msg, logs

        def on_ok(result):
            ok, msg, logs = result
            for l in logs:
                self.main.log(l)
            if ok:
                if self.sync_check.isChecked() and self.main.xlsx_path:
                    try:
                        sync_ok, sync_msg = sync_delete_person(self.main.xlsx_path, name)
                        self.main.log(sync_msg)
                    except Exception as e:
                        self.main.log(f"Excel同步失败: {e}")
                QMessageBox.information(self, "成功", f"批量迁出成功\n{msg}")
                self._clear_results()
                self.name_edit.clear()
            else:
                QMessageBox.warning(self, "失败", msg)
            self.delete_btn.setEnabled(True)

        self.main.run_worker(do_delete, on_ok)

    def _delete_by_plate(self):
        if not self._found_credentials:
            return
        plate = self.plate_edit.text().strip()
        name = self.name_edit.text().strip()

        confirm = QMessageBox.question(
            self, "确认删除",
            f"将执行以下操作:\n"
            f"1. 注销车牌: {plate}\n"
            f"2. 删除用户: {name}（批量迁出，同时删除关联凭证）\n\n"
            f"此操作不可撤销，确定继续？",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return

        self.delete_btn.setEnabled(False)
        self.main.log(f"注销车牌并删除用户: {plate} - {name}")

        def do_delete():
            api = self.main.api
            logs = []

            # 步骤1: 注销凭证
            logs.append("[1/3] 注销车牌凭证...")
            active_cred_ids = [c["id"] for c in self._found_credentials if c.get("status") == 1]
            skipped = len(self._found_credentials) - len(active_cred_ids)
            if active_cred_ids:
                ok, msg = api.cancel_credential(active_cred_ids)
                logs.append(f"  {msg}")
                if not ok:
                    return False, msg, logs
            if skipped:
                logs.append(f"  已跳过 {skipped} 条已注销凭证")
            logs.append("[1/3] 凭证注销完成")

            # 步骤2: 查找用户并批量迁出
            logs.append("[2/3] 查找用户...")
            persons = api.query_person(name)
            if not persons:
                logs.append("[2/3] 未找到用户，跳过用户删除")
                return True, "车牌注销成功，但未找到关联用户", logs

            person_ids = [p["id"] for p in persons]
            logs.append(f"[2/3] 找到 {len(persons)} 个用户，执行批量迁出...")
            ok, msg = api.batch_migrate_out(person_ids)
            logs.append(f"[2/3] {msg}")

            # 步骤3: 同步Excel
            logs.append("[3/3] 同步本地表格...")
            return ok, msg, logs

        def on_ok(result):
            ok, msg, logs = result
            for l in logs:
                self.main.log(l)
            if ok:
                if self.sync_check.isChecked() and self.main.xlsx_path and name:
                    try:
                        sync_ok, sync_msg = sync_delete_person(self.main.xlsx_path, name)
                        self.main.log(sync_msg)
                    except Exception as e:
                        self.main.log(f"Excel同步失败: {e}")
                QMessageBox.information(self, "成功", "车牌注销并删除用户完成")
                self._clear_results()
                self.plate_edit.clear()
                self.name_edit.clear()
            else:
                QMessageBox.warning(self, "失败", msg)
            self.delete_btn.setEnabled(True)

        self.main.run_worker(do_delete, on_ok)


# ==================== 黑白名单Tab ====================

class BlackWhiteListTab(QWidget):
    def __init__(self, main: MainWindow):
        super().__init__()
        self.main = main
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        # 操作区
        op_group = QGroupBox("名单操作")
        op_layout = QFormLayout(op_group)

        self.type_combo = QComboBox()
        self.type_combo.addItems(LIST_TYPE_MAP.keys())

        self.plate_edit = QLineEdit()
        self.plate_edit.setPlaceholderText("例: 粤S12345")

        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDisplayFormat("yyyy-MM-dd")
        self.start_date.setDate(QDate.currentDate())

        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDisplayFormat("yyyy-MM-dd")
        self.end_date.setDate(QDate.currentDate().addDays(3))

        self.remark_edit = QLineEdit()
        self.remark_edit.setPlaceholderText("选填")

        op_layout.addRow("名单类型:", self.type_combo)
        op_layout.addRow("车  牌:", self.plate_edit)
        op_layout.addRow("开始日期:", self.start_date)
        op_layout.addRow("结束日期:", self.end_date)
        op_layout.addRow("备  注:", self.remark_edit)
        layout.addWidget(op_group)

        # 按钮行
        btn_row = QHBoxLayout()
        add_btn = QPushButton("添加名单")
        add_btn.clicked.connect(self._on_add)
        query_btn = QPushButton("查询名单")
        query_btn.clicked.connect(self._on_query)
        del_btn = QPushButton("删除选中")
        del_btn.clicked.connect(self._on_delete)
        btn_row.addWidget(add_btn)
        btn_row.addWidget(query_btn)
        btn_row.addWidget(del_btn)
        layout.addLayout(btn_row)

        # 查询过滤
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("查询类型:"))
        self.query_type_combo = QComboBox()
        self.query_type_combo.addItems(["全部", "黑名单", "灰名单", "白名单"])
        filter_row.addWidget(self.query_type_combo)
        self.query_plate_edit = QLineEdit()
        self.query_plate_edit.setPlaceholderText("车牌筛选(可选)")
        filter_row.addWidget(self.query_plate_edit)
        layout.addLayout(filter_row)

        # 结果表格
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["ID", "车牌", "类型", "有效期", "备注"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.table)

    def _type_name(self, t):
        for name, val in LIST_TYPE_MAP.items():
            if val == t:
                return name
        return str(t)

    def _on_add(self):
        plate = self.plate_edit.text().strip()
        if not plate:
            QMessageBox.warning(self, "提示", "请输入车牌号")
            return

        list_type = LIST_TYPE_MAP[self.type_combo.currentText()]
        start = self.start_date.date().toString("yyyy-MM-dd")
        end = self.end_date.date().toString("yyyy-MM-dd")
        remark = self.remark_edit.text().strip()
        type_name = self.type_combo.currentText()

        self.main.log(f"添加{type_name}: {plate}")

        def do_add():
            return self.main.api.add_bw_record(plate, list_type, start, end, remark)

        def on_ok(result):
            ok, msg = result
            self.main.log(msg)
            if ok:
                QMessageBox.information(self, "成功", f"{type_name}添加成功")
                self._on_query()
            else:
                QMessageBox.warning(self, "失败", msg)

        self.main.run_worker(do_add, on_ok)

    def _on_query(self):
        plate = self.query_plate_edit.text().strip()
        type_text = self.query_type_combo.currentText()
        list_type = LIST_TYPE_MAP.get(type_text)
        self.main.log(f"查询名单: 类型={type_text}, 车牌={plate or '全部'}")

        def do_query():
            return self.main.api.query_bw_list(plate, list_type)

        def on_ok(rows):
            self.table.setRowCount(len(rows))
            for i, row in enumerate(rows):
                self.table.setItem(i, 0, QTableWidgetItem(str(row.get("id", ""))))
                self.table.setItem(i, 1, QTableWidgetItem(row.get("credentialNo", "")))
                self.table.setItem(i, 2, QTableWidgetItem(
                    self._type_name(row.get("listType", 0))
                ))
                start = row.get("startTime", "")[:10]
                end = row.get("endTime", "")[:10]
                self.table.setItem(i, 3, QTableWidgetItem(f"{start} ~ {end}"))
                self.table.setItem(i, 4, QTableWidgetItem(row.get("remark", "")))
            self.main.log(f"查询到 {len(rows)} 条记录")

        self.main.run_worker(do_query, on_ok)

    def _on_delete(self):
        rows = self.table.selectionModel().selectedRows()
        if not rows:
            QMessageBox.warning(self, "提示", "请先选择要删除的记录")
            return

        ids = []
        for idx in rows:
            rid = self.table.item(idx.row(), 0).text()
            plate = self.table.item(idx.row(), 1).text()
            ids.append((rid, plate))

        confirm = QMessageBox.question(
            self, "确认删除",
            f"确定删除 {len(ids)} 条记录？",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return

        for rid, plate in ids:
            self.main.log(f"删除记录: {plate} (ID: {rid})")

            def make_del(rid_, plate_):
                def do_del():
                    return self.main.api.delete_bw_record(rid_)
                return do_del

            def make_on(plate_):
                def on_ok(result):
                    ok, msg = result
                    self.main.log(f"{plate_}: {msg}")
                return on_ok

            self.main.run_worker(make_del(rid, plate), make_on(plate))

        # 刷新
        from PyQt5.QtCore import QTimer
        QTimer.singleShot(1000, self._on_query)


# ==================== 入口 ====================

def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    # 设置中文字体
    font = app.font()
    font.setFamily("Microsoft YaHei")
    font.setPointSize(10)
    app.setFont(font)

    dialog = LoginDialog()
    if dialog.exec_() != QDialog.Accepted:
        sys.exit(0)

    window = MainWindow(dialog.api, dialog.xlsx_path)
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
